"""
Parse 10 đề thi vào 10 (Tiếng Anh) docx -> 1 sheet Excel gộp Exam + Question.
"""

from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EXAM_DIR = Path(r"c:\Users\tungd\Downloads\10 đề thi vào 10")
OUTPUT_FILE = Path(r"D:\TaO10-BackEnd\data\import_10_de_thi_vao_10.xlsx")

W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

SECTION_PATTERNS = [
    (
        r"Mark the letter A, B, C, or D on your answer sheet to indicate the word whose underlined part differs from the other three in pronunciation",
        "Phát âm",
    ),
    (
        r"Mark the letter A, B, C, or D on your answer sheet to indicate the word that differs from the other three in the position of primary stress",
        "Trọng âm",
    ),
    (
        r"Mark the letter A, B, C, or D on your answer sheet to indicate the underlined part that needs correction",
        "Tìm lỗi sai",
    ),
    (
        r"Mark the letter A, B, C, or D on your answer sheet to indicate the correct answer to each of the following questions",
        "Chọn đáp án đúng",
    ),
    (
        r"Mark the letter A, B, C, or D on your answer sheet to indicate the correct response to each of the following exchanges",
        "Giao tiếp",
    ),
    (
        r"Mark the letter A, B, C, or D on your answer sheet to indicate the word\(s\) CLOSEST in meaning",
        "Từ đồng nghĩa",
    ),
    (
        r"Mark the letter A, B, C, or D on your answer sheet to indicate the word\(s\) OPPOSITE in meaning",
        "Từ trái nghĩa",
    ),
    (
        r"Read the following passage.*?numbered blanks",
        "Điền từ vào đoạn văn",
    ),
    (
        r"Read the following passage.*?each of the questions",
        "Đọc hiểu",
    ),
    (
        r"Mark the letter A, B, C or D on your answer sheet to indicate the sentence that is closest in meaning",
        "Viết lại câu (gần nghĩa)",
    ),
    (
        r"Mark the letter A, B, C or D to indicate the sentence that is best written from the words/ phrases given",
        "Viết lại câu (từ cho sẵn)",
    ),
]

QUESTION_RE = re.compile(
    r"Câu\s+(\d+)\.\s*(.*?)\s*A\.\s*(.*?)\s*B\.\s*(.*?)\s*C\.\s*(.*?)\s*D\.\s*(.*?)"
    r"(?=Câu\s+\d+\.|Mark the letter|Read the following passage|ĐÁP ÁN|$)",
    re.DOTALL | re.IGNORECASE,
)

CLOZE_PASSAGE_RE = re.compile(
    r"Read the following passage.*?numbered blanks\.(.*?)(?=Câu\s+25\.)",
    re.DOTALL | re.IGNORECASE,
)

READING_PASSAGE_RE = re.compile(
    r"Read the following passage.*?each of the questions\.(.*?)(?=Câu\s+31\.)",
    re.DOTALL | re.IGNORECASE,
)

ANSWER_RE = re.compile(r"[A-D]", re.IGNORECASE)

HEADERS = [
    "Title",
    "Description",
    "QuestionsCount",
    "DurationTime",
    "Level",
    "Year",
    "ExamType",
    "ViewsCount",
    "AttemptsCount",
    "QuestionNumber",
    "Section",
    "QuestionText",
    "OptionA",
    "OptionB",
    "OptionC",
    "OptionD",
    "CorrectAnswer",
    "Explanation",
    "Points",
]

PASSAGE_QUESTION_NUMBERS = {
    "cloze": 24,
    "reading": 30,
}

NA = "N/A"


def extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    parts: list[str] = []
    for node in root.iter(f"{W_NS}t"):
        if node.text:
            parts.append(node.text)
        if node.tail:
            parts.append(node.tail)
    return "".join(parts)


def normalize_whitespace(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def parse_answers(text: str) -> dict[int, str]:
    marker = re.search(r"ĐÁP\s*ÁN", text, re.IGNORECASE)
    if not marker:
        raise ValueError("Không tìm thấy phần ĐÁP ÁN")

    tail = text[marker.end() :]
    letters = ANSWER_RE.findall(tail)
    if len(letters) < 40:
        raise ValueError(f"Chỉ tìm thấy {len(letters)} đáp án, cần 40")

    return {i + 1: letters[i].upper() for i in range(40)}


def detect_sections(text: str) -> list[tuple[int, str, str]]:
    sections: list[tuple[int, str, str]] = []
    for pattern, name in SECTION_PATTERNS:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            sections.append((match.start(), name, match.group(0)))
    sections.sort(key=lambda x: x[0])
    return sections


def section_for_question(sections: list[tuple[int, str, str]], pos: int) -> tuple[str, str]:
    current_name = "Khác"
    current_instruction = ""
    for start, name, instruction in sections:
        if start <= pos:
            current_name = name
            current_instruction = instruction
        else:
            break
    return current_name, current_instruction


def extract_exam_meta(text: str, file_name: str) -> dict:
    title_match = re.search(r"ĐỀ\s+ÔN\s+THI\s+VÀO\s+LỚP\s+10[-–]?\s*ĐỀ\s*(\d+)", text, re.IGNORECASE)
    exam_no = title_match.group(1) if title_match else re.search(r"De-(\d+)", file_name, re.IGNORECASE).group(1)

    return {
        "exam_no": int(exam_no),
        "title": f"Đề ôn thi vào lớp 10 - Tiếng Anh - Đề {exam_no}",
        "description": (
            "Đề thi thử vào lớp 10 môn Tiếng Anh (form mới). "
            "Nguồn: thuvienhoclieu.com"
        ),
    }


def extract_passage(pattern: re.Pattern[str], text: str) -> str:
    match = pattern.search(text)
    if not match:
        return ""
    return normalize_whitespace(match.group(1))


def build_explanation(q: dict) -> str:
    if q.get("is_passage"):
        section = q.get("section", "")
        if section == "Điền từ vào đoạn văn":
            return (
                "Đây là bài đọc dùng cho phần điền từ (câu 25-30). "
                "Học sinh đọc đoạn văn trước khi trả lời các câu hỏi bên dưới."
            )
        return (
            "Đây là bài đọc dùng cho phần đọc hiểu (câu 31-34). "
            "Học sinh đọc đoạn văn trước khi trả lời các câu hỏi bên dưới."
        )

    correct = q.get("correct_answer", "")
    options = {
        "A": q.get("option_a", ""),
        "B": q.get("option_b", ""),
        "C": q.get("option_c", ""),
        "D": q.get("option_d", ""),
    }
    section = q.get("section", "")

    if not correct:
        return "Không có đáp án cho câu hỏi này."

    correct_text = options.get(correct, "")
    base = f"Đáp án đúng là {correct}"
    if correct_text:
        base += f" ({correct_text})"

    section_hints = {
        "Phát âm": ". Chọn từ có phần gạch chân phát âm khác các từ còn lại.",
        "Trọng âm": ". Chọn từ có trọng âm rơi vào vị trí khác các từ còn lại.",
        "Tìm lỗi sai": ". Chọn phần gạch chân sai về ngữ pháp hoặc cách dùng từ.",
        "Chọn đáp án đúng": ". Chọn đáp án đúng theo ngữ pháp và ngữ cảnh.",
        "Giao tiếp": ". Chọn câu trả lời phù hợp nhất trong tình huống giao tiếp.",
        "Từ đồng nghĩa": ". Chọn từ/cụm từ gần nghĩa nhất với từ được gạch chân.",
        "Từ trái nghĩa": ". Chọn từ/cụm từ trái nghĩa với từ được gạch chân.",
        "Điền từ vào đoạn văn": ". Chọn từ/cụm từ phù hợp nhất để điền vào chỗ trống.",
        "Đọc hiểu": ". Dựa vào nội dung bài đọc để chọn đáp án đúng.",
        "Viết lại câu (gần nghĩa)": ". Chọn câu có nghĩa gần giống nhất với câu gốc.",
        "Viết lại câu (từ cho sẵn)": ". Sắp xếp các từ cho sẵn thành câu hoàn chỉnh đúng ngữ pháp.",
    }
    return base + section_hints.get(section, ".")


def parse_questions(text: str) -> list[dict]:
    answer_pos = re.search(r"ĐÁP\s*ÁN", text, re.IGNORECASE)
    question_text = text[: answer_pos.start()] if answer_pos else text

    sections = detect_sections(question_text)
    answers = parse_answers(text)
    by_number: dict[int, dict] = {}

    for match in QUESTION_RE.finditer(question_text):
        q_num = int(match.group(1))
        q_body = normalize_whitespace(match.group(2))
        section_name, section_instruction = section_for_question(sections, match.start())

        if not q_body:
            q_body = section_instruction

        by_number[q_num] = {
            "question_number": q_num,
            "section": section_name,
            "question_text": q_body,
            "option_a": normalize_whitespace(match.group(3)),
            "option_b": normalize_whitespace(match.group(4)),
            "option_c": normalize_whitespace(match.group(5)),
            "option_d": normalize_whitespace(match.group(6)),
            "correct_answer": answers.get(q_num, ""),
            "is_passage": False,
        }

    if len(by_number) != 40:
        raise ValueError(f"Parsed {len(by_number)} questions, expected 40")

    cloze_passage = extract_passage(CLOZE_PASSAGE_RE, question_text)
    reading_passage = extract_passage(READING_PASSAGE_RE, question_text)

    ordered: list[dict] = []
    for n in range(1, 25):
        ordered.append(by_number[n])

    if cloze_passage:
        ordered.append(
            {
                "question_number": PASSAGE_QUESTION_NUMBERS["cloze"],
                "section": "Điền từ vào đoạn văn",
                "question_text": cloze_passage,
                "option_a": NA,
                "option_b": NA,
                "option_c": NA,
                "option_d": NA,
                "correct_answer": NA,
                "is_passage": True,
            }
        )

    for n in range(25, 31):
        ordered.append(by_number[n])

    if reading_passage:
        ordered.append(
            {
                "question_number": PASSAGE_QUESTION_NUMBERS["reading"],
                "section": "Đọc hiểu",
                "question_text": reading_passage,
                "option_a": NA,
                "option_b": NA,
                "option_c": NA,
                "option_d": NA,
                "correct_answer": NA,
                "is_passage": True,
            }
        )

    for n in range(31, 41):
        ordered.append(by_number[n])

    for q in ordered:
        q["explanation"] = build_explanation(q)
        q["points"] = 0 if q.get("is_passage") else 0.25

    return ordered


def style_header(ws, headers: list[str]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = max(14, min(50, len(header) + 4))


def non_null(value) -> str | int | float:
    if value is None:
        return NA
    if isinstance(value, str) and not value.strip():
        return NA
    return value


def write_row(ws, row: int, exam: dict, q: dict) -> None:
    values = [
        exam["title"],
        exam["description"],
        40,
        60,
        "Lớp 9",
        2026,
        "THI_THU",
        0,
        0,
        q["question_number"],
        q["section"],
        q["question_text"],
        q["option_a"],
        q["option_b"],
        q["option_c"],
        q["option_d"],
        q["correct_answer"],
        q["explanation"],
        q["points"],
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(row=row, column=col, value=non_null(value))
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    docx_files = sorted(EXAM_DIR.glob("*.docx"), key=lambda p: int(re.search(r"De-(\d+)", p.name, re.I).group(1)))
    if len(docx_files) != 10:
        raise SystemExit(f"Expected 10 docx files, found {len(docx_files)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    style_header(ws, HEADERS)

    row = 2
    for docx_path in docx_files:
        raw = extract_docx_text(docx_path)
        meta = extract_exam_meta(raw, docx_path.name)
        questions = parse_questions(raw)

        for q in questions:
            write_row(ws, row, meta, q)
            ws.row_dimensions[row].height = 30
            row += 1

        print(f"OK: {docx_path.name} ({len(questions)} rows)")

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")
    print(f"Total rows: {row - 2}")


if __name__ == "__main__":
    main()
